"""
Export utilities: PDF (ReportLab) and Excel (openpyxl).
Usage: call generate_guest_pdf / generate_guest_excel with the list of guests.
"""
import io
from typing import List
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ─── PDF ──────────────────────────────────────────────────────────────────────

def generate_guest_pdf(guests: List[dict], wedding_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"{wedding_name} — Guest List", styles["Title"]))
    elements.append(Spacer(1, 12))

    headers = ["Name", "Phone", "Side", "RSVP", "Pax", "Table", "Meal Pref"]
    data = [headers]
    for g in guests:
        data.append([
            g.get("name", ""),
            g.get("phone", ""),
            g.get("side", ""),
            g.get("rsvp_status", ""),
            str(g.get("pax_count", 1)),
            str(g.get("table_number", "")),
            g.get("meal_preference", ""),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f3ff")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def generate_budget_pdf(items: List[dict], wedding_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"{wedding_name} — Budget Summary", styles["Title"]))
    elements.append(Spacer(1, 12))

    headers = ["Category", "Item", "Estimated (RM)", "Actual (RM)", "Paid (RM)", "Balance (RM)", "Due Date"]
    data = [headers]
    for item in items:
        estimated = float(item.get("estimated_amount") or 0)
        actual = float(item.get("actual_amount") or 0)
        paid = float(item.get("paid_amount") or 0)
        balance = actual - paid
        data.append([
            item.get("category", ""),
            item.get("item_name", ""),
            f"{estimated:,.2f}",
            f"{actual:,.2f}",
            f"{paid:,.2f}",
            f"{balance:,.2f}",
            str(item.get("payment_due_date", "")),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


# ─── Excel ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _apply_header(ws, headers: List[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def generate_guest_excel(guests: List[dict], wedding_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Guest List"
    headers = ["Name", "Phone", "Email", "Side", "RSVP Status", "Pax", "Table", "Meal Preference", "Allergies", "VIP", "Notes"]
    _apply_header(ws, headers)

    for row, g in enumerate(guests, 2):
        ws.cell(row=row, column=1, value=g.get("name", ""))
        ws.cell(row=row, column=2, value=g.get("phone", ""))
        ws.cell(row=row, column=3, value=g.get("email", ""))
        ws.cell(row=row, column=4, value=g.get("side", ""))
        ws.cell(row=row, column=5, value=g.get("rsvp_status", ""))
        ws.cell(row=row, column=6, value=g.get("pax_count", 1))
        ws.cell(row=row, column=7, value=g.get("table_number", ""))
        ws.cell(row=row, column=8, value=g.get("meal_preference", ""))
        ws.cell(row=row, column=9, value=g.get("allergies", ""))
        ws.cell(row=row, column=10, value="Yes" if g.get("is_vip") else "No")
        ws.cell(row=row, column=11, value=g.get("notes", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_budget_excel(items: List[dict], wedding_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    headers = ["Category", "Item", "Estimated (RM)", "Actual (RM)", "Paid (RM)", "Balance (RM)", "Due Date", "Notes"]
    _apply_header(ws, headers)

    for row, item in enumerate(items, 2):
        estimated = float(item.get("estimated_amount") or 0)
        actual = float(item.get("actual_amount") or 0)
        paid = float(item.get("paid_amount") or 0)
        ws.cell(row=row, column=1, value=item.get("category", ""))
        ws.cell(row=row, column=2, value=item.get("item_name", ""))
        ws.cell(row=row, column=3, value=estimated)
        ws.cell(row=row, column=4, value=actual)
        ws.cell(row=row, column=5, value=paid)
        ws.cell(row=row, column=6, value=actual - paid)
        ws.cell(row=row, column=7, value=str(item.get("payment_due_date", "")))
        ws.cell(row=row, column=8, value=item.get("notes", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
